from openpyxl import load_workbook

import random
import time

wb = load_workbook("vokabeln.xlsx")

sheet_names = wb.sheetnames

dict_sheets = {}
counter = 1
sheet = None

for sheet_name in sheet_names:
    dict_sheets[counter] = sheet_name
    counter += 1

while sheet is None:
    for key in dict_sheets:
        print(str(key) + ":", dict_sheets[key])
    try:
        sheet = dict_sheets[int(input("Nummer der Liste eingeben: "))]
    except KeyError:
        pass
    except ValueError:
        pass
    except KeyboardInterrupt:
        pass

sheet_file = wb[sheet]

swedish = []
german = []

for cell in sheet_file:
    swedish.append(cell[0].value)
    german.append(cell[1].value)

while True:
    word = random.randint(0, len(swedish) - 1)
    print(swedish[word])
    if input("Deutsch: ").lower() == german[word].lower():
        print("Richtig!")
    else:
        print("Falsch, richtig gewesen wäre: ", german[word])
    time.sleep(3)
